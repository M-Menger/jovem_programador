import openpyxl
from datetime import datetime

dbMultas = 'C:/Users/Balcão/Desktop/EQS_Engenharia/4.Multas/DBMultas.xlsx'

def searchAIT(auto):
    dataBase = openpyxl.load_workbook(dbMultas)
    mySheetdb = dataBase['DBMultas']

    for row in range(1, 2200):
        if mySheetdb[f'P{row}'].value == auto:
            return row
    return None

def pegaDados(numCell):
    dataBase = openpyxl.load_workbook(dbMultas, data_only=True)
    mySheetdb = dataBase['DBMultas']

    idCar = mySheetdb[f'K{numCell}'].value
    data = mySheetdb[f'O{numCell}'].value
    cod = mySheetdb[f'Y{numCell}'].value
    cep = mySheetdb[f'W{numCell}'].value
    city = mySheetdb[f'X{numCell}'].value
    infr = mySheetdb[f'Z{numCell}'].value

    return idCar, data, cod, infr, cep, city

def pegaValor(cod):
    database = openpyxl.load_workbook(dbMultas)
    mySheetdb = database['CTB']

    for row in range(1, 450):
        if mySheetdb[f'A{row}'].value == cod:
            return mySheetdb[f'C{row}'].value
    return None

def pegaDestino(placa):
    dataBase = openpyxl.load_workbook(dbMultas)
    mySheetdb = dataBase['Destinatario']

    for row in range(1, 2200):
        if mySheetdb[f'A{row}'].value == placa:
            return mySheetdb[f'C{row}'].value
    return None

def endProccess(data, auto):
    dataBase = openpyxl.load_workbook(dbMultas)
    mySheetdb = dataBase['LogEnvio']

    row = 2
    while mySheetdb[f'A{row}'].value is not None:
        row += 1

    mySheetdb[f'B{row}'] = data
    mySheetdb[f'A{row}'] = auto

    dataBase.save(dbMultas)
